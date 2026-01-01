from openpyxl import load_workbook
import sys
import json

excel_path = sys.argv[1]
payload = json.loads(sys.argv[2])

name = payload.get("name", "")
org = payload.get("org", "")
grade = payload.get("grade", "")
date = payload.get("date", "")
answers = payload.get("answers", [])

wb = load_workbook(excel_path, data_only=True)

# Dashboard
dashboard = wb["Dashboard"]
start_row = 8
start_col = 3

for i, val in enumerate(answers):
    dashboard.cell(row=start_row, column=start_col + i, value=val)

# Report
report = wb["Report"]
report["A6"] = name
report["B6"] = grade
report["A7"] = org
report["B7"] = date

# --- Extract RIASEC scores ---
# --- Extract RIASEC scores ---
riasec_rows = range(51, 57)
riasec_scores = {}

for row in riasec_rows:
    letter = report[f"A{row}"].value
    score = report[f"B{row}"].value

    if letter is None:
        continue

    try:
        score = int(score)
    except (TypeError, ValueError):
        score = 0  # fallback safety

    riasec_scores[str(letter).lower()] = score

# --- Extract personality analyzer value ---
personality_analyzer = report["A14"].value

# --- Extract SECE scores ---
sece = wb["SECE"]

sece_scores = {}

for row in range(4, 20):  # B4:C19
    score = sece[f"B{row}"].value
    letter = sece[f"C{row}"].value

    if letter is None:
        continue

    try:
        score = int(score)
    except (TypeError, ValueError):
        score = 0

    sece_scores[str(letter).lower()] = score

# Sort SECE scores
sorted_sece = sorted(sece_scores.items(), key=lambda x: x[1], reverse=True)

top_2 = [item[0] for item in sorted_sece[:2]]
bottom_2 = [item[0] for item in sorted_sece[-2:]]

# --- Print result as JSON for Node.js ---
result = {
    "riasec": riasec_scores,
    "personalityAnalyzer": personality_analyzer,
    "sece": {
        "all": sece_scores,
        "top2": top_2,
        "bottom2": bottom_2
    }
}

print(json.dumps(result))

wb.save(excel_path)
