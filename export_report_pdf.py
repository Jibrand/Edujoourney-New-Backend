from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
import sys
import os
import subprocess

excel_path = sys.argv[1]
report_sheet_name = sys.argv[2]

base, _ = os.path.splitext(excel_path)
pdf_path = base + ".pdf"
fixed_pdf_path = base + "_final.pdf"

# ---------- EXCEL PREP ----------
wb = load_workbook(excel_path)

if report_sheet_name not in wb.sheetnames:
    raise Exception(f"Sheet '{report_sheet_name}' not found")

report_ws = wb[report_sheet_name]

# Hide all other sheets (keep formulas working)
for sheet in wb.worksheets:
    if sheet.title != report_sheet_name:
        sheet.sheet_state = "hidden"

# Set print area (adjust if needed)
report_ws.print_title_rows = None
report_ws.print_title_cols = None
report_ws.print_area = "A1:Z50"

# Activate report sheet
wb.active = wb.index(report_ws)
wb.save(excel_path)

# ---------- EXCEL → PDF ----------
subprocess.run([
    "soffice",
    "--headless",
    "--convert-to", "pdf",
    excel_path,
    "--outdir", os.path.dirname(excel_path)
], check=True)

# ---------- REMOVE LAST PAGE ----------
reader = PdfReader(pdf_path)
writer = PdfWriter()

# Keep all pages EXCEPT the last one
for i in range(len(reader.pages) - 1):
    writer.add_page(reader.pages[i])

with open(fixed_pdf_path, "wb") as f:
    writer.write(f)

# Replace original PDF
os.remove(pdf_path)
os.rename(fixed_pdf_path, pdf_path)
